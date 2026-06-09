"""
report_generator.py

Generate Excel MIS report.

Author: Sagar
"""

from datetime import datetime
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

def format_headers(worksheet) -> None:
    """
    Apply standard formatting
    to worksheet headers.
    """

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    worksheet.freeze_panes = "A2"


def auto_adjust_columns(worksheet) -> None:
    """
    Automatically resize columns.
    """

    for column in worksheet.columns:

        max_length = 0

        column_letter = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except Exception:
                pass

        worksheet.column_dimensions[
            column_letter
        ].width = max_length + 3

def create_executive_summary_sheet(
    workbook,
    metrics
):
    """
    Create executive summary dashboard.
    """

    ws = workbook.active
    ws.title = "Executive Summary"

    sales = metrics["sales"]

    total_revenue = sales["net_revenue"].sum()

    stockout_count = len(
        metrics["inventory_health"]
    )

    top_region = (
        metrics["region_revenue"]
        .iloc[0]["region"]
    )

    total_sales = len(sales)

    total_returns = len(
        metrics["return_rate"]
    )

    return_rate = round(
        (total_returns / total_sales) * 100,
        2
    )

    # Report Title

    ws.merge_cells("A1:B1")

    ws["A1"] = (
        "RETAILCORE MIS DASHBOARD"
    )

    ws["A1"].font = Font(
        size=16,
        bold=True
    )

    ws["A1"].alignment = Alignment(
        horizontal="center"
    )

    # Report Date

    ws["A3"] = "Report Date"
    ws["B3"] = datetime.now().strftime(
        "%d-%b-%Y"
    )

    # KPI Table

    ws["A5"] = "Metric"
    ws["B5"] = "Value"

    ws["A6"] = "Total Revenue"
    ws["B6"] = total_revenue

    ws["A7"] = "Return Rate (%)"
    ws["B7"] = return_rate

    ws["A8"] = "Top Region"
    ws["B8"] = top_region

    ws["A9"] = "Stockout Products"
    ws["B9"] = stockout_count

    for cell in ws[5]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ws["B6"].number_format = (
        '₹#,##0.00'
    )

    # Add Weekly Revenue Chart

    weekly_chart = Image(
        "charts/weekly_revenue_trend.png"
    )

    weekly_chart.width = 500
    weekly_chart.height = 250

    ws.add_image(
        weekly_chart,
        "D2"
    )

    auto_adjust_columns(ws)

    
def create_sales_analysis_sheet(
    workbook,
    metrics
):
    """
    Sales analysis using
    Region x Category pivot.
    """

    ws = workbook.create_sheet(
        "Sales Analysis"
    )

    sales = metrics["sales"]

    pivot = sales.pivot_table(
        index="region",
        columns="category",
        values="net_revenue",
        aggfunc="sum",
        fill_value=0
    )

    ws.append(
        ["Region"] +
        list(pivot.columns)
    )

    for region, row in zip(
        pivot.index,
        pivot.values.tolist()
    ):
        ws.append(
            [region] + row
        )

    format_headers(ws)

    auto_adjust_columns(ws)

def create_inventory_alert_sheet(
    workbook,
    metrics
):
    """
    Inventory shortage report.
    """

    ws = workbook.create_sheet(
        "Inventory Alerts"
    )

    inventory = (
        metrics["inventory_health"]
        .copy()
    )

    inventory["shortage"] = (
        inventory["reorder_level"]
        -
        inventory["stock_available"]
    )

    inventory = (
        inventory
        .sort_values(
            "shortage",
            ascending=False
        )
    )

    ws.append(
        inventory.columns.tolist()
    )

    for row in (
        inventory.values.tolist()
    ):
        ws.append(row)

    format_headers(ws)

    auto_adjust_columns(ws)


def create_returns_sheet(
    workbook,
    metrics
):
    """
    Return rate analysis.
    """

    ws = workbook.create_sheet(
        "Returns Deep Dive"
    )

    returns = metrics[
        "return_rate"
    ]

    ws.append(
        returns.columns.tolist()
    )

    for row in returns.values.tolist():
        ws.append(row)

    format_headers(ws)

    auto_adjust_columns(ws)

def create_charts_sheet(
    workbook
):
    """
    Add charts to workbook.
    """

    ws = workbook.create_sheet(
        "Charts"
    )

    # Weekly Revenue Chart

    weekly_chart = Image(
        "charts/weekly_revenue_trend.png"
    )

    weekly_chart.width = 700
    weekly_chart.height = 300

    ws.add_image(
        weekly_chart,
        "A1"
    )

    # Region Revenue Chart

    region_chart = Image(
        "charts/region_revenue.png"
    )

    region_chart.width = 700
    region_chart.height = 300

    ws.add_image(
        region_chart,
        "A22"
    )

    # Give rows some height
    for row in range(1, 50):
        ws.row_dimensions[row].height = 20


def generate_excel_report(
    metrics,output_dir
):
    """
    Generate complete MIS report.
    """

    workbook = Workbook()

    create_executive_summary_sheet(
        workbook,
        metrics
    )

    create_sales_analysis_sheet(
        workbook,
        metrics
    )

    create_inventory_alert_sheet(
        workbook,
        metrics
    )

    create_returns_sheet(
        workbook,
        metrics
    )

    create_charts_sheet(
        workbook
    )

    os.makedirs(
        output_dir,
        exist_ok=True
    )

    output_dir = Path(output_dir)

    report_name = (
        output_dir
        /
        f"MIS_Report_"
        f"{datetime.now():%Y%m%d}"
        f".xlsx"
    )

    workbook.save(
        report_name
    )

    return str(
        report_name
    )